#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Oct 30 00:10:24 2020

@author: hugokleikamp
"""


#%% clear variables

# try:
#     from IPython import get_ipython
#     get_ipython().magic('clear')
#     get_ipython().magic('reset -f')
# except:
#     pass

#%% change directory to script directory (should work on windows and mac)
import os
from pathlib import Path
from inspect import getsourcefile
os.chdir(str(Path(os.path.abspath(getsourcefile(lambda:0))).parents[0]))
print(os.getcwd())

#%%
import urllib
import zipfile
from openpyxl import Workbook, load_workbook 
import datetime
import pandas as pd


#%% Krona template

if "krona_template.xlsm" not in os.listdir():
    print("Downloading Krona file")
    Kronaurl="https://github.com/marbl/Krona/releases/download/xl2.5/Krona.xltm.zip"
    urllib.request.urlretrieve(Kronaurl, "Krona.xltm.zip")
    with zipfile.ZipFile("Krona.xltm.zip", 'r') as zip_ref:
        print("Unzipping Krona file")
        zip_ref.extractall()
        
            
    wb = load_workbook('Krona.xltm', read_only=False,keep_vba=True)
    wb.template = False
    wb.save('Krona.xlsm')
    
    wb = load_workbook('Krona.xlsm', keep_vba=True)
    ws=wb.active
    for row in ws['A4:F19']:
      for cell in row:
        cell.value = ""
    
    ws['C3']=""
    ws['D2']="lineage"
    ws['B3']="frequency"
    
    wb.save('krona_template.xlsm')

#%% KEGG KO file

#download
print("Downloading Kegg file")
KEGGurl="https://www.genome.jp/kegg-bin/download_htext?htext=ko00001&format=htext&filedir="
urllib.request.urlretrieve(KEGGurl, "ko00001.keg")

#read and parse
with open("ko00001.keg") as file:
    print("Parsing Kegg file")
    lines=file.readlines()

unnested=list()
columns=["A","B","C","D"]
for line in lines:
    if line.startswith("A"): 
        cat1=line[1:].strip() 
        unnested.append("/t".join([cat1]))
    if line.startswith("B"): 
        cat2=line[1:].strip()
        unnested.append("/t".join([cat1,cat2]))
    if line.startswith("C"): 
        cat3=line[1:].strip()
        unnested.append("/t".join([cat1,cat2,cat3]))
    if line.startswith("D"): 
        cat4=line[1:].strip()
        unnested.append("/t".join([cat1,cat2,cat3,cat4]))


df=pd.DataFrame(unnested,columns=["val"])
df=df["val"].str.rsplit("/t",expand=True)
df=df.iloc[:,: 4]
df.columns=["cat1","cat2","cat3","cat4"]
df=df[~df["cat4"].isnull()]

#split cat4
df[["KO","Genes"]] = df["cat4"].str.split(" ", 1, expand=True)
df["Description"]=df["Genes"].str.split("; ", 1, expand=True)[1]
df["Genes"]=df["Genes"].str.split("; ", 1, expand=True)[0]
df["ec"]=df["Description"].str.split("EC:", 1, expand=True)[1].str[:-1]
df["Description"]=df["Description"].str.split("EC:", 1, expand=True)[0].str[:-1]

#explode genes and EC nos for exact string matching
df["Genes"]=df["Genes"].str.split(",")
df=df.explode('Genes').reset_index(drop=True)

df["ec"]=df["ec"].str.split(" ")
df=df.explode("ec").reset_index(drop=True)

# strip spaces
for i in df.columns:
    df[i]=df[i].astype(str).str.strip()

# save as tsp
df.to_csv(str(datetime.datetime.today()).split()[0]+"keg.tsv",sep="\t")
