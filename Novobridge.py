#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov  28 17:32:32 2020

@author: hugokleikamp
"""
#%% clear variables and console

try:
    from IPython import get_ipython
    get_ipython().magic('clear')
    get_ipython().magic('reset -f')
except:
    pass

"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
                                                        "Parameters & setup" 
"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

#%% parameters Part 1: Unipept submission
starting_vars=dir()

#filter parameters
ALC_cutoff=40      # miniumum required ALC score (Peaks score)
Score_cutoff=-0.1  # mininum required score cutoff (DeepNovo score)
ppm_cutoff=20      # maximum allowed ppm
length_cutoff=7    # mininum required peptide length
Area_cutoff=0      # mininum required peak area
Intensity_cutoff=0 # mininum required intensity

#get variables for writing to output
current_vars=set(dir()).difference(starting_vars)
parameters=[i for i in locals().items() if i[0] in current_vars]

#%% parameters Part 2: Composition 
starting_vars=dir()

#filter parameters
comp_ALC_cutoff=70      # miniumum required ALC score (Peaks score)
comp_Score_cutoff=-0.1 # mininum required score cutoff (DeepNovo score)
comp_ppm_cutoff=15      # maximum allowed ppm
comp_length_cutoff=7    # mininum required peptide length
comp_Area_cutoff=0      # mininum required peak area
comp_Intensity_cutoff=0 # mininum required intensity
cutbranch=3    # minimum amount of unique peptides per taxonomic branch in denoising

#Which taxa to annotate
comp_ranks=["superkingdom","phylum","class","order","family","genus","species"]

#quantification parameters
tax_count_targets=["Spectral_counts"]#,"Area","Intensity"]
tax_count_methods=["average","total","topx"] 
tax_topx=5
tax_normalize=False # normalize quantification to total for that rank

#get variables for writing to output
current_vars=set(dir()).difference(starting_vars)
comp_parameters=[i for i in locals().items() if i[0] in current_vars]

#%% Parameters Part 3: Function
starting_vars=dir()

#which pathways to include
Pathways=['09100 Metabolism', 
          '09120 Genetic Information Processing'
          '09130 Environmental Information Processing'
          '09140 Cellular Processes'] 

#which levels to include
cats=["cat1","cat2","cat3","cat4"]                                    

#quantification parameters
fun_count_targets=["Spectral_counts"]#,"Area","Intensity"]
fun_count_methods=["average","total","topx"] 
fun_topx=5
fun_normalize=False # normalize quantification to total for that rank

current_vars=set(dir()).difference(starting_vars)
fun_parameters=comp_parameters+[i for i in locals().items() if i[0] in current_vars]

#%% Modules
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

import random, re, requests
import threading, time, string


from itertools import chain
from collections import Counter
from openpyxl import load_workbook 



#%% change directory to script directory (should work on windows and mac)
import os
from pathlib import Path
from inspect import getsourcefile
os.chdir(str(Path(os.path.abspath(getsourcefile(lambda:0))).parents[0]))
print(os.getcwd())

#%% Functions

# make stacked bars

def stacked_bar(taxa,df,ylabel,pathout,filename): 

    labels=[i.split("_name")[0] for i in taxa];
    countcols=[i for i in df.columns if "count" in i]
    absdat=df[countcols] 
    absdat.columns=labels
    normdat= absdat/np.nansum(absdat.to_numpy(),axis=0)
    
    figure, axes = plt.subplots(1, 2)
    ax1=absdat.T.plot(ax=axes[0],kind='bar', stacked=True, figsize=(10, 6), legend=False)
    ax1.set_ylabel(ylabel)
    ax1.set_xlabel('taxonomic ranks')
    ax1.set_xticklabels(labels, rotation=30)
    ax1.set_title("Absolute")
    
    ax2=normdat.T.plot(ax=axes[1],kind='bar', stacked=True, figsize=(10, 6), legend=False)
    ax2.set_ylabel(ylabel)
    ax2.set_xlabel('taxonomic ranks')
    ax2.set_xticklabels(labels, rotation=30)
    ax2.set_title("Normalized")
    
    
    plt.gcf().suptitle(Path(basename).stem)
    
    figname=str(Path(pathout,(filename.replace(os.path.splitext(filename)[1], '.png'))))
    
    plt.savefig(figname)
    
    #plot legend in separate plot
    
    
    
    return figname    



#mass calculation for DeepNovo                    
std_aa_mass = {'G': 57.02146, 'A': 71.03711, 'S': 87.03203, 'P': 97.05276, 'V': 99.06841,
               'T': 101.04768,'C': 103.00919,'L': 113.08406,'I': 113.08406,'J': 113.08406,
               'N': 114.04293,'D': 115.02694,'Q': 128.05858,'K': 128.09496,'E': 129.04259,
               'M': 131.04049,'H': 137.05891,'F': 147.06841,'U': 150.95364,'R': 156.10111,
               'Y': 163.06333,'W': 186.07931,'O': 237.14773}  

def mass_calc(x,std_aa_mass=std_aa_mass):
    return sum(std_aa_mass.get(aa) for aa in x)+18.01056

# Threading unipept
def unipept_scrape(r,url):
    
    while True:
        try:
            r.extend(requests.get(url,stream=True).json())
            break
        except:
            "sleeping"
            time.sleep(2)
            
            
# chunker
def chunks(lst,n):
    for i in range(0,len(lst),n):
        yield lst[i:i+n]


#%% 

"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
                                                    "Part 1: Annotation with unipept" 
"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

pathin="input_peaks" #input folder
s=time.time()
for filename in os.listdir(pathin): 
    if Path(filename).stem[0].isalnum(): # check if not a hidden file, filename should start with alphanumeric
        for Randomize in [False]: #, 'scramble']: 
            xlsdf=[]
            
            if filename.endswith('.txt'):                        
                with open(filename,"r") as f: 
                    xlsdf=pd.DataFrame(f.readlines())
            
            #added dynamic delimiter detection
            if filename.endswith('.csv'):
                xlsdf=pd.read_csv(str(Path(pathin,filename)),sep=",")
                if "Peptide" not in xlsdf.columns: 
                    delims=[i[0] for i in Counter([i for i in str(xlsdf.iloc[0]) if not i.isalnum()]).most_common()]
                    for delim in delims:
                        if delim==" ": sep="\s"
                        try:
                            xlsdf=pd.read_csv(str(Path(pathin,filename)),sep=delim)
                            if "Peptide" in xlsdf.columns:
                                break
                        except:
                            pass
                        
            if filename.endswith('.tsv'):                               xlsdf=pd.read_csv(str(Path(pathin,filename)),sep="\t")
            if filename.endswith('.xlsx') or filename.endswith('.xls'): xlsdf=pd.read_excel(str(Path(pathin,filename)),engine='openpyxl')   
            if len(xlsdf):
                xlsdf=xlsdf.fillna("0") #replace nans with zero
                
                #%% if DeepNovo output, convert to PEAKS output
                if 'predicted_sequence' in xlsdf.columns:
                    
                    xlsdf["Peptide"]=xlsdf["predicted_sequence"]           
                    xlsdf=xlsdf[xlsdf["Peptide"]!="0"]
                
                    #calculate ppm shift from m/z
                    mass=np.zeros((1,len(xlsdf)))
                    mass+=xlsdf["Peptide"].str.count("(Carbamidomethylation)").fillna(0)*57.021463
                    mass+=xlsdf["Peptide"].str.count("(Oxidation)").fillna(0)*15.994915
    
                    xlsdf['Peptide']=xlsdf['Peptide'].apply(lambda x: re.sub("[\(\[].*?[\)\]]", "", x).replace(",","")) #remove ptms in peptides
                    
    
                    
                    #%% calculate peptide mass foor DeepNovo  
                    
                    xlsdf['calculated_mass']=mass[0]+xlsdf['Peptide'].apply(lambda x: mass_calc(x)).values
                    xlsdf['precursor_mass']=xlsdf['precursor_mz']*xlsdf['precursor_charge']-xlsdf['precursor_charge']*1.007277                
                    xlsdf["ppm"]=(1000000/xlsdf['calculated_mass'])*(xlsdf['calculated_mass']-xlsdf['precursor_mass'])
    
                    #rename columns
                    if "feature_id" in xlsdf.columns: xlsdf=xlsdf.rename(columns={"feature_id":"Scan"})  
                    if "feature_area" in xlsdf.columns: xlsdf=xlsdf.rename(columns={"feature_area":"Area"})  
                    if "feature_intensity" in xlsdf.columns: xlsdf=xlsdf.rename(columns={"feature_intensity":"Intensity"})  
                
                #set datatypes as float
                xlsdf["Tag Length"]=xlsdf['Peptide'].apply(len)
                xlsdf=xlsdf.fillna("0") #replace nans with zero
                
                for i in ['Tag Length','ALC (%)','predicted_score','ppm','Area','Intensity']:
                    if i in xlsdf.columns:
                        xlsdf[i]=xlsdf[i].astype(float) 
                
                #add Scan if not present, add Scan
                if 'Scan' not in xlsdf.columns: 
                    xlsdf['Scan']=list(range(0,len(xlsdf)))
                            
                # recalibrate ppm   
                if 'ppm' in xlsdf.columns:
                    hist, bin_edges = np.histogram(xlsdf['ppm'].tolist(),bins=100)
                    xlsdf['ppm']=abs(xlsdf['ppm'] -(bin_edges[np.where(hist==np.max(hist))[0]]
                                                   +bin_edges[np.where(hist==np.max(hist))[0]+1])/2)
                #remove ptms in peptides
                peplist=xlsdf['Peptide'].tolist()
                xlsdf['Peptide']=list(map(lambda x: re.sub("[\(\[].*?[\)\]]", "", x),peplist)) 
                
                #%% filter by ALC scores, ppm and peptide lengths 
                filt="on"
                if filt=="on":    
                    if 'Tag Length' in xlsdf.columns: 
                        xlsdf=xlsdf[xlsdf['Tag Length']>=length_cutoff]
                    
                    if 'ALC (%)' in xlsdf.columns:
                        xlsdf=xlsdf[xlsdf['ALC (%)']>=ALC_cutoff] #scoring Peaks 
                    
                    if 'predicted_score' in xlsdf.columns:
                        xlsdf=xlsdf[xlsdf['predicted_score']<=Score_cutoff] #scoring DeepNovo 
                    
                    if 'ppm' in xlsdf.columns: 
                        xlsdf=xlsdf[xlsdf['ppm']<=ppm_cutoff]
                    
                    if 'Area' in xlsdf.columns: 
                        xlsdf=xlsdf[xlsdf['Area']>=Area_cutoff]
                
                    if 'Intensity' in xlsdf.columns: 
                        xlsdf=xlsdf[xlsdf['Intensity']>=Intensity_cutoff]
                        
                #%% randomization (optional)
                if   Randomize=='scramble': #scramble all in front of cleavage site
                     xlsdf['Peptide']=[''.join(random.sample(i[:len(i)-1], len(i)-1)+[i[-1]]) for i in xlsdf['Peptide']]
                
                #%% if multiple candidates, sort by scan, add candidate list and select best
                if 'ALC (%)' in xlsdf.columns: xlsdf=xlsdf.sort_values(['Scan', 'ALC (%)','Tag Length'], ascending=[1, 0, 0])
                elif 'predicted_score' in xlsdf.columns: xlsdf=xlsdf.sort_values(['Scan', 'predicted_score','Tag Length'], ascending=[1, 0, 0])
                else: xlsdf=xlsdf.sort_values(['Scan','Tag Length'], ascending=[1, 0, 0])
                xlsdf['Candidate'] = xlsdf.groupby('Scan').cumcount() + 1
                xlsdf=xlsdf[xlsdf["Candidate"]==1]
                        
                #%% submit peptides to unipept
                
                #urls
                twl='http://api.unipept.ugent.be/api/v1/pept2lca.json?input[]=';
                twr='&equate_il=true&extra=true&names=true';
                fwl='http://api.unipept.ugent.be/api/v1/pept2funct.json?input[]=';
                fwr='&equate_il=true';
                
                ranks=[rank+"_name" for rank in comp_ranks]
                fields=["peptide"]+["taxon_name"]+ranks
                
                base_thread=threading.active_count()
                unipeps=np.unique(xlsdf['Peptide'])
                batchsize=100
                steps=list(range(0,len(unipeps),batchsize))
                taxalist=list()
                funlist=list()
    
                threads=[]
                counter=0
                for chunk in chunks(unipeps,batchsize):
                    counter+=1
                    print(counter)
                    query="&input[]=".join(chunk)
                        
                    #taxonomy
                    turl=twl+query+twr 
                    t=threading.Thread(target=unipept_scrape, args=[taxalist,turl])
                    t.start()
                    threads.append(t)
                    
                    #function
                    furl=fwl+query+fwr 
                    t=threading.Thread(target=unipept_scrape, args=[funlist,furl])
                    t.start()
                    threads.append(t)
                    
                    #unwind in case of thread overload, manage server traffic
                    cur_thread=threading.active_count()
                    if (cur_thread-base_thread)>200:
                        print("unwinding, query at: "+str(counter/len(unipeps)))
                        for thread in threads:
                            thread.join()
                        threads=[] #this seems to act different on windows?
                     
                for thread in threads:
                    thread.join()
                
                #%% post processing
                #remove non-regular taxonomic ranks
                taxdf=pd.DataFrame(taxalist)
                [taxdf.pop(x) for x in taxdf.columns if x not in fields]
                
                #parse function dataframe
                funs=pd.DataFrame(funlist)
                funs["ec"]= funs["ec"].apply( lambda x: " ".join(pd.json_normalize(x)["ec_number"]) if x else [])
                funs["go"]= funs["go"].apply( lambda x: " ".join(pd.json_normalize(x)["go_term"]) if x else [])
                funs["ipr"]=funs["ipr"].apply(lambda x: " ".join(pd.json_normalize(x)["code"]) if x else [])
    
                #merge annotations
                taxdf.set_index('peptide',inplace=True, drop=True)
                funs.set_index('peptide',inplace=True, drop=True)
                xlsdf.set_index('Peptide',inplace=True, drop=False)
                xlsdf=xlsdf.join(taxdf,how='left')
                xlsdf=xlsdf.join(funs,how='left')
                
                xlsdf=xlsdf.mask(xlsdf.applymap(str).eq('[]'))
                xlsdf=xlsdf.fillna("")
                #%% write result
    

    
                pathout="output_unipept"
                if not os.path.exists(pathout): os.makedirs(pathout)
                
                basename="unipept_nb_"+Path(filename).stem+ '.tsv'
                if Randomize=="scramble": basename="Rand_"+basename
                outfilename=str(Path(pathout,basename))
                xlsdf.to_csv(outfilename,sep="\t")


        
        #%%
                """""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
                                                                    "Part 2: Compositional analysis" 
                """""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
                
                
                #%% Pre_processing filtering & denoising
                
                #if multiple candidates, pick best
                
                
                
                #%%
                
                filt="on"
                if filt=="on":    
                    if 'Tag Length' in xlsdf.columns: 
                        xlsdf=xlsdf[xlsdf['Tag Length']>=comp_length_cutoff]
                    
                    if 'ALC (%)' in xlsdf.columns:
                        xlsdf=xlsdf[xlsdf['ALC (%)']>=comp_ALC_cutoff]
                        
                    if 'predicted_score' in xlsdf.columns:
                        xlsdf=xlsdf[xlsdf['predicted_score']<=comp_Score_cutoff] #scoring DeepNovo 
                    
                    if 'ppm' in xlsdf.columns: 
                        xlsdf=xlsdf[xlsdf['ppm']<=comp_ppm_cutoff]
                    
                    if 'Area' in xlsdf.columns: 
                        xlsdf=xlsdf[xlsdf['Area']>=comp_Area_cutoff]
                
                    if 'Intensity' in xlsdf.columns: 
                        xlsdf=xlsdf[xlsdf['Intensity']>=comp_Intensity_cutoff]
                
                  
                denoise="on"
                if denoise=="on":
                    unirows=xlsdf[["Peptide"]+ranks].drop_duplicates()[ranks].astype(str).values.tolist()
                    jbranch=["#*".join(i) for i in unirows]
                    fbranch=[branch for branch, counts in Counter(jbranch).items() if counts >= cutbranch]
                    allowed_taxids=set(chain(*[i.split("#*") for i in fbranch]))
                    for i in ranks:
                        xlsdf.loc[~xlsdf[i].isin(allowed_taxids),i]="" 
                            
                #%% krona plot (only coded for spectral counting)
                
                pathout="output_composition"
                if not os.path.exists(pathout): os.makedirs(pathout)
                
                if "krona_template.xlsm" in os.listdir():
                    
                    grouped=xlsdf.groupby(ranks)
                    vals=grouped.size() 
                    vals=vals.reset_index(name="Count")  
                    vals=vals[~(vals[ranks]=="").all(axis=1)] #remove empty rows
                    
                    #results
                    branches=vals[ranks].values.tolist()  
                    counts=  vals["Count"].tolist()
                    
                    #fill gaps
                    for i in range(len(branches)):
                        j=[j for j in range(len(branches[i])) if branches[i][j]!=""]
                        for l in range(0,max(j)):
                            if branches[i][l]=="": branches[i][l]="annotation_gap"     
                    vals[ranks]=branches
                    
                    branchdf=vals
    
                    kronafilename=str(Path(pathout,"Spectral_counts_Krona_"+Path(filename).stem+ '.xlsm'))
                    letters=list(string.ascii_uppercase)
                    wb = load_workbook("krona_template.xlsm",read_only=False, keep_vba=True)
                    ws = wb.active
                    for i in range(0,len(branches)):
                    
                        ws['A{0}'.format(4+i)].value=filename
                        ws['B{0}'.format(4+i)].value=counts[i]
                        for j in range(0,len(ranks)):
                            ws['{0}{1}'.format(letters[j+3],4+i)].value=branches[i][j]
                    wb.save(kronafilename)   
                    
                else:
                    print("No Krona template found, proceding without generating krona plots")
                    
  
                
                #%% quantification and visual outputs, write output to xlsx
                pathout="output_composition"
                if not os.path.exists(pathout): os.makedirs(pathout)
                

      
                if type(tax_count_targets)==str: tax_count_targets=list(tax_count_targets)
                for target in tax_count_targets:
                    
                        
                    if target!="Spectral_counts" and target not in xlsdf.columns:
                        print("Target column: '"+str(target)+"' not found in data, please change parameter: 'count_targets'")
                        continue
                        
                    if type(tax_count_methods)==str: tax_count_methods=list(tax_count_methods)
                    for method in tax_count_methods:
                        
            
            
                        quantdf=pd.DataFrame()
                        for rank in ranks:
                            
                            if target=="Spectral_counts":          
                                values=Counter(xlsdf[rank].astype(str))
                                if "" in values.keys(): values.pop("") #ignore unassigned peptides
                                
                            else:
                                xlsdf[target]=xlsdf[target].astype(float)
                                grouped=xlsdf.groupby(rank)[target]
                                if method=="average": values=grouped.mean()
                                elif method=="total":values=grouped.sum()
                                elif method=="topx": values=grouped.nlargest(tax_topx).sum(level=0)
                    
                            values=pd.Series(values).sort_values(axis=0, ascending=False, inplace=False).reset_index()
                            values.columns=[rank,rank+"_count"]   
                            values=values[values[rank]!=""]
                            if tax_normalize==True: 
                                values[rank+"_count"]=values[rank+"_count"]/values[rank+"_count"].sum()*100 #normalize to 100%
                            
                            quantdf = pd.concat([quantdf, values], axis=1) 
                
                        #writing output
                        pathout="output_composition"
                        if not os.path.exists(pathout): os.makedirs(pathout)
                    
                        quantdfs=quantdf.fillna(0)
                        namecols=[i for i in quantdf.columns  if "count" not in i] 
                        countcols=[i for i in quantdf.columns  if "count" in i] 
                       
                        
                        basename=Path(filename).stem+ '.xlsx'
                        if method=="topx": method=method.replace("x",str(fun_topx))
                        if target=='Spectral_counts': basename="composition_"+target+"_"+basename
                        else: basename="composition_"+method+"_"+target+"_"+basename
                        if Randomize=="scramble": basename="Rand_"+basename
                        xlsfilename=str(Path(pathout,basename))
                        
                        writer = pd.ExcelWriter(xlsfilename, engine='xlsxwriter')
                        
                        pardf=pd.DataFrame(comp_parameters,columns=["Name","Value"])
                        pardf.loc[pardf["Name"]=="tax_count_targets","Value"]=target
                        pardf.loc[pardf["Name"]=="tax_count_methods","Value"]=method
                        if target=='Spectral_counts': pardf=pardf[pardf["Name"]!="tax_count_methods"]
                        pd.DataFrame(pardf.to_excel(writer, sheet_name='Parameters'))
                              
                        quantdf[namecols].to_excel(writer, sheet_name='TAX_LINEAGES')
                        quantdf[countcols].to_excel(writer, sheet_name='TAX_COUNTS')
                        
                        if target=='Spectral_counts': stacked_bar(ranks,quantdf,target,pathout, basename)
                        else:                         stacked_bar(ranks,quantdf,method+"_"+target,pathout, basename)
                        
 
                        writer.save()   


                
                                        

                #%%
        
                """""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
                                                                    "Part 3: Functional analysis" 
                """""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
    
                #%% pre processing
                
                #preselect xlsdf to only have rows that have ec numbers
                xlsdf=xlsdf[xlsdf["ec"].notnull()]
                #calculte total area/intensity/spectral counts for normalization

                #separate ec annotations for exact matching
                xlsdf["ec"]=xlsdf["ec"].apply(lambda x: str(x).split(" ")) 
               
                #%% quantification
                
                #check if kegg database is present
                if not any("keg.tsv" in i for i in os.listdir()):
                    print("no local kegg database found, please run download_utilities.py before functional annotation can be done")
                    
                else: #Load local kegg database
                    for tsv in os.listdir():
                        if tsv.endswith('keg.tsv') and tsv[0].isalnum():  
                            keggdf  = pd.read_csv(tsv,sep="\t")

                    
                    #only select pathways that are in the parameters
                    keggdf=keggdf[keggdf.isin(Pathways).any(axis=1)] 
                    #only select cats that are in the parameters

                    if "ec" not in cats: keggdf=keggdf.loc[keggdf.isin(Pathways).any(axis=1),cats+["ec"]] 
                    else:                 keggdf=keggdf.loc[keggdf.isin(Pathways).any(axis=1),cats] 
                    
                    if type(fun_count_targets)==str: tax_count_methods=list(fun_count_targets)
                    for target in fun_count_targets:
                        
                        if target!="Spectral_counts" and target not in xlsdf.columns:
                            print("Target column: '"+str(target)+"' not found in data, please change parameter: 'count_targets'")
                            continue
                        
                        if type(fun_count_methods)==str: fun_count_methods=list(fun_count_methods)
                        for method in fun_count_methods:
                            

                            #match pathways to annotations and explode
                            if target=="Spectral_counts":             fundf=xlsdf[["Scan","ec"]]
                            else:                                     fundf=xlsdf[["Scan","ec",target]]
                            
                            fundf=fundf.explode("ec")
                            fundf=fundf[fundf["ec"].notnull()]
                            fundf=fundf[fundf["ec"]!=""]
                            fundf=fundf.merge(keggdf,how="left",on="ec")    
                            
                            quantdf=fundf[cats].drop_duplicates()
                            for cat in cats:
                                
                                if target=="Spectral_counts":             catdf=fundf[["Scan",cat]].drop_duplicates()
                                else: catdf=fundf[["Scan",cat,target]].drop_duplicates()
                                catdf=catdf.explode(cat)
                                
                                if target=="Spectral_counts":
                                    values=pd.Series(Counter(catdf[cat].astype(str))).rename_axis(cat)
                                    
                                else:
                                    catdf[target]=catdf[target].astype(float)
                                    grouped=catdf.groupby(cat)[target]
                                    if method=="average": values=grouped.mean()
                                    elif method=="total": values=grouped.sum()
                                    elif method=="topx":  values=grouped.nlargest(fun_topx).sum(level=0)
                                    
                                                                  
                                values=values.reset_index(drop=False)
                                values.columns=[cat,cat+"_"+target]                        
                                if fun_normalize==True: values[cat,"cat_"+target]=values[cat,cat+"_"+target]/sum(values)       
                                quantdf=quantdf.merge(values,how="left",on=cat)    
       
                                
                            quantdf=quantdf.dropna()
                            #writing output
                            pathout="output_function"
                            if not os.path.exists(pathout): os.makedirs(pathout)
                            
                            basename=Path(filename).stem+ '.xlsx'
                            if method=="topx": method=method.replace("x",str(fun_topx))
                            if target=='Spectral_counts': basename="function_"+target+"_"+basename
                            else: basename="function_"+method+"_"+target+"_"+basename
                            if Randomize=="scramble": basename="Rand_"+basename
                            xlsfilename=str(Path(pathout,basename))
                            
                            
                            writer = pd.ExcelWriter(xlsfilename, engine='xlsxwriter')
                            
                            
                            pardf=pd.DataFrame(fun_parameters,columns=["Name","Value"])
                            pardf.loc[pardf["Name"]=="fun_count_targets","Value"]=target
                            pardf.loc[pardf["Name"]=="fun_count_methods","Value"]=method
                            if target=='Spectral_counts': pardf=pardf[pardf["Name"]!="fun_count_methods"]
                            pd.DataFrame(pardf.to_excel(writer, sheet_name='Parameters'))
                            
                
                            for i in cats:
                                cols=[j for j in quantdf.columns if i in j]
                                #add previous cols
                                c=0; prev=list()
                                while i!=cats[c]:
                                    prev.append(cats[c])
                                    c+=1
                                cols=prev+cols
                                quantdf[cols].drop_duplicates().to_excel(writer, sheet_name=i) 
                             
                            writer.save()   
                 
                        
#%% Cleanup (remove .png files)
cleanup=True
if cleanup:
    [os.remove(i) for i in os.listdir() if i.endswith(".png") and i[0].isalnum()]

print("elaped time: "+str(time.time()-s))



